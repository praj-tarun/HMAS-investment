"""
holdings_importer.py — Import holdings from broker xlsx files into portfolio memory.

Supports two file types:
  1. Stock broker statement (Zerodha/Kite style)
     Headers: Stock Name | ISIN | Quantity | Average buy price | ...
  2. MF consolidated statement (CAMS/KFintech style)
     Headers: Scheme Name | AMC | Category | ... | Units | Invested Value | Current Value | ...

Usage:
    from src.data.holdings_importer import sync_holdings_to_memory
    from src.memory.memory_mcp import MemoryMCP
    memory = MemoryMCP()
    report = sync_holdings_to_memory(memory, folder="holdings")
"""

from __future__ import annotations

import os
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

# ---------------------------------------------------------------------------
# ISIN -> NSE ticker map (ETFs only — stocks are not mapped here)
# ---------------------------------------------------------------------------

ISIN_TO_TICKER: Dict[str, str] = {
    # Gold ETFs
    "INF204KB17I5": "GOLDBEES.NS",   # Nippon India ETF Gold BeES
    "INF204KB15I8": "GOLDBEES.NS",   # Nippon India ETF Gold BeES (older series)
    "INF200K01UB2": "GOLDBEES.NS",
    "INF205K01DN2": "HDFCGOLD.NS",   # HDFC Gold ETF
    "INF179KB1HL0": "HDFCGOLD.NS",
    "INF082J01306": "ICICIGOLD.NS",  # ICICI Pru Gold ETF
    "INF732E01500": "KOTAKGOLD.NS",  # Kotak Gold ETF

    # Silver ETFs
    "INF204KB19H2": "SILVERBEES.NS", # Nippon India ETF Silver BeES
    "INF205K01FP4": "HDFCSILVER.NS", # HDFC Silver ETF

    # Nifty 50 ETFs
    "INF204KB16G5": "NIFTYBEES.NS",  # Nippon India ETF Nifty BeES
    "INF082J01282": "ICICINIFTY.NS", # ICICI Pru Nifty ETF
    "INF179K01YS6": "HDFCNIFTY.NS",  # HDFC Nifty ETF
    "INF200K01YM3": "KOTAKNIFTY.NS", # Kotak Nifty ETF
    "INF732E01369": "KOTAKNIFTY.NS",

    # Nifty Next 50 / Junior BeES
    "INF204KB13N0": "JUNIORBEES.NS", # Nippon India ETF Junior BeES
    "INF082J01332": "ICICISENEX.NS", # ICICI Pru Nifty Next 50

    # Sensex ETFs
    "INF200K01VA5": "SENSEXETF.NS",
    "INF205K01DM4": "HDFCSENSEX.NS",

    # Bank Nifty ETFs
    "INF204KB11D4": "BANKBEES.NS",   # Nippon India ETF Bank BeES
    "INF082J01217": "ICICIBANKN.NS",

    # Midcap ETFs (MIDCPNIFTY.NS / KOTAKMIDCP.NS not available on Yahoo Finance)
    # Using MIDSMALL.NS (Mirae MidSmallcap400) as the closest working proxy
    "INF204KB14N8": "MIDSMALL.NS",
    "INF732E01773": "MIDSMALL.NS",

    # IT / Sectoral ETFs
    "INF204KB16N6": "ITBEES.NS",     # Nippon India ETF IT BeES
    "INF082J01308": "ICICIIT.NS",

    # PSU Bank ETF
    "INF204KB12N2": "PSUBNKBEES.NS", # Nippon India ETF PSU Bank BeES

    # Liquid / Overnight
    "INF204KB16H3": "LIQUIDBEES.NS", # Nippon India ETF Liquid BeES
}


# ---------------------------------------------------------------------------
# MF category -> proxy ETF for technical analysis
# ---------------------------------------------------------------------------

MF_PROXY_MAP: Dict[str, Tuple[str, str]] = {
    # (ticker, description)
    # Mid cap: MIDSMALL.NS = Mirae Asset Nifty MidSmallcap400 Momentum ETF
    "mid cap":  ("MIDSMALL.NS", "Mirae MidSmallcap400 ETF proxy (midcap direction)"),
    "midcap":   ("MIDSMALL.NS", "Mirae MidSmallcap400 ETF proxy (midcap direction)"),
    # Small cap: JUNIORBEES.NS = Nippon India ETF Nifty Next 50
    "small cap": ("JUNIORBEES.NS", "Nifty Junior BeES proxy (small/mid blend)"),
    "smallcap":  ("JUNIORBEES.NS", "Nifty Junior BeES proxy (small/mid blend)"),
    # Infra: INFRABEES.NS = Nippon India ETF Infra BeES
    "infra":          ("INFRABEES.NS", "Nippon India Infra BeES proxy"),
    "infrastructure": ("INFRABEES.NS", "Nippon India Infra BeES proxy"),
    # Large/Flexi/Multi cap: NIFTYBEES.NS = Nippon India ETF Nifty BeES
    "large cap": ("NIFTYBEES.NS", "Nifty 50 ETF proxy"),
    "largecap":  ("NIFTYBEES.NS", "Nifty 50 ETF proxy"),
    "flexi cap": ("NIFTYBEES.NS", "Nifty 50 ETF proxy"),
    "multi cap": ("NIFTYBEES.NS", "Nifty 50 ETF proxy"),
    # Sectoral
    "banking": ("BANKBEES.NS", "Bank Nifty BeES proxy"),
    "bank":    ("BANKBEES.NS", "Bank Nifty BeES proxy"),
    "it":      ("ITBEES.NS",   "Nippon India IT BeES proxy"),
    "pharma":  ("PHARMABEES.NS", "Pharma ETF proxy"),
}


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class StockHolding:
    name: str
    isin: str
    ticker: Optional[str]          # NSE ticker if resolved from ISIN
    quantity: float
    avg_buy_price: float
    current_price: float
    unrealized_pnl: float
    unrealized_pnl_pct: float


@dataclass
class MFHolding:
    scheme_name: str
    amc: str
    category: str
    sub_category: str
    units: float
    invested_value: float
    current_value: float
    returns: float                 # absolute Rs.
    xirr: float                    # annualised %
    proxy_ticker: Optional[str]    # suggested ETF proxy
    proxy_note: str
    is_equity: bool


@dataclass
class ImportReport:
    stocks: List[StockHolding]
    mfs: List[MFHolding]
    synced_tickers: List[str]
    skipped_isins: List[str]       # stock ISINs with no ticker mapping
    mf_proxies_added: List[str]    # MF proxy tickers added as portfolio holdings
    mf_debt_skipped: List[str]     # debt MF schemes skipped
    errors: List[str]


# ---------------------------------------------------------------------------
# File detection
# ---------------------------------------------------------------------------

def detect_file_type(filepath: Path) -> str:
    """
    Returns 'stock', 'mf', or 'unknown'.
    Scans every cell in the first 30 rows × 15 cols looking for key headers.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        text = set()
        for row in ws.iter_rows(max_row=30, max_col=15, values_only=True):
            for cell in row:
                if cell and isinstance(cell, str):
                    text.add(cell.strip().lower())
        wb.close()
        if "isin" in text and ("stock name" in text or "scrip name" in text):
            return "stock"
        if "scheme name" in text and "amc" in text:
            return "mf"
        return "unknown"
    except Exception:
        return "unknown"


# ---------------------------------------------------------------------------
# Header row finder
# ---------------------------------------------------------------------------

def _find_header_row(ws, expected_cols: List[str], max_rows: int = 40) -> Optional[int]:
    """
    Return the 1-based row index of the header row, or None if not found.
    Matches if ≥ len(expected_cols)//2 + 1 of the expected columns appear.
    """
    threshold = max(1, len(expected_cols) // 2 + 1)
    for row_idx, row in enumerate(ws.iter_rows(max_row=max_rows, values_only=True), start=1):
        row_lower = {str(c).strip().lower() for c in row if c is not None}
        matches = sum(1 for col in expected_cols if col.lower() in row_lower)
        if matches >= threshold:
            return row_idx
    return None


def _row_to_dict(header_row: tuple, data_row: tuple) -> Dict[str, str]:
    """Map header names -> cell values for a data row."""
    result: Dict[str, str] = {}
    for h, v in zip(header_row, data_row):
        if h is not None:
            result[str(h).strip().lower()] = v
    return result


# ---------------------------------------------------------------------------
# Stock holdings parser
# ---------------------------------------------------------------------------

def parse_stock_holdings(filepath: Path) -> List[StockHolding]:
    """
    Parse a stock broker holdings statement xlsx.
    Expected columns: Stock Name, ISIN, Quantity, Average buy price,
                      Closing price, Unrealised P&L
    """
    import openpyxl

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    expected = ["stock name", "isin", "quantity", "average buy price", "closing price"]
    header_row_idx = _find_header_row(ws, expected)
    if header_row_idx is None:
        wb.close()
        raise ValueError(f"Could not find header row in {filepath.name}. "
                         f"Expected columns: {expected}")

    # Re-read to get all rows as lists
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    headers = all_rows[header_row_idx - 1]
    holdings: List[StockHolding] = []

    for data_row in all_rows[header_row_idx:]:
        row = _row_to_dict(headers, data_row)

        name = str(row.get("stock name") or row.get("scrip name") or "").strip()
        isin = str(row.get("isin") or "").strip()

        if not name or not isin or isin.upper() == "ISIN":
            continue                          # skip blank / sub-header rows
        if len(isin) != 12:
            continue                          # not a real ISIN

        def _num(key: str) -> float:
            for k in [key, key.replace(" ", ""), key.replace(" buy ", " ")]:
                v = row.get(k)
                if v is not None:
                    try:
                        return float(str(v).replace(",", ""))
                    except (ValueError, TypeError):
                        pass
            return 0.0

        qty = _num("quantity")
        avg_buy = _num("average buy price")
        closing = _num("closing price")
        unrealised_pnl = _num("unrealised p&l") or _num("unrealized p&l") or _num("p&l")

        if qty == 0 and avg_buy == 0:
            continue

        pnl_pct = ((closing - avg_buy) / avg_buy * 100) if avg_buy else 0.0
        ticker = ISIN_TO_TICKER.get(isin)

        holdings.append(StockHolding(
            name=name,
            isin=isin,
            ticker=ticker,
            quantity=qty,
            avg_buy_price=avg_buy,
            current_price=closing,
            unrealized_pnl=unrealised_pnl,
            unrealized_pnl_pct=pnl_pct,
        ))

    return holdings


# ---------------------------------------------------------------------------
# MF holdings parser
# ---------------------------------------------------------------------------

def parse_mf_holdings(filepath: Path) -> List[MFHolding]:
    """
    Parse a MF consolidated account statement xlsx.
    Expected columns: Scheme Name, AMC, Category, Sub-category,
                      Units, Invested Value, Current Value, Returns, XIRR
    """
    import openpyxl

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    # Try "Holdings" sheet first, then fall back to active
    ws = wb["Holdings"] if "Holdings" in wb.sheetnames else wb.active

    expected = ["scheme name", "amc", "category", "units", "invested value", "current value"]
    header_row_idx = _find_header_row(ws, expected)
    if header_row_idx is None:
        wb.close()
        raise ValueError(f"Could not find header row in {filepath.name}. "
                         f"Expected columns: {expected}")

    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    headers = all_rows[header_row_idx - 1]
    holdings: List[MFHolding] = []

    for data_row in all_rows[header_row_idx:]:
        row = _row_to_dict(headers, data_row)

        scheme = str(row.get("scheme name") or "").strip()
        amc = str(row.get("amc") or "").strip()
        category = str(row.get("category") or "").strip()
        sub_cat = str(row.get("sub-category") or row.get("sub category") or "").strip()

        if not scheme or scheme.lower() == "scheme name":
            continue
        if not amc or len(amc) < 2:
            continue

        def _num(key: str) -> float:
            v = row.get(key)
            if v is not None:
                try:
                    return float(str(v).replace(",", "").replace("%", ""))
                except (ValueError, TypeError):
                    pass
            return 0.0

        units = _num("units")
        invested = _num("invested value")
        current = _num("current value")
        returns = _num("returns")
        xirr = _num("xirr")

        # Determine if equity
        cat_lower = category.lower()
        is_equity = any(k in cat_lower for k in [
            "equity", "mid cap", "small cap", "large cap", "multi cap", "flexi cap",
            "infrastructure", "banking", "pharma", "it", "sectoral", "thematic",
            "elss", "tax saving",
        ])
        is_debt = any(k in cat_lower for k in [
            "debt", "liquid", "overnight", "money market", "short term",
            "ultra short", "low duration", "credit risk", "gilt", "banking and psu",
        ])
        if is_debt:
            is_equity = False

        # Find proxy ticker
        proxy_ticker = None
        proxy_note = ""
        full_text = (cat_lower + " " + sub_cat.lower() + " " + scheme.lower())
        for key, (ticker, note) in MF_PROXY_MAP.items():
            if key in full_text:
                proxy_ticker = ticker
                proxy_note = note
                break

        holdings.append(MFHolding(
            scheme_name=scheme,
            amc=amc,
            category=category,
            sub_category=sub_cat,
            units=units,
            invested_value=invested,
            current_value=current,
            returns=returns,
            xirr=xirr,
            proxy_ticker=proxy_ticker,
            proxy_note=proxy_note,
            is_equity=is_equity,
        ))

    return holdings


# ---------------------------------------------------------------------------
# Main sync function
# ---------------------------------------------------------------------------

def sync_holdings_to_memory(memory, folder: str = "holdings") -> ImportReport:
    """
    Scan *folder* for xlsx files, parse each, and upsert into portfolio memory.

    - Stock ETF holdings with a known ISIN -> ticker mapping are added as
      portfolio holdings using the actual avg buy price and closing price.
    - Equity MF holdings are added as portfolio holdings using the proxy ETF
      ticker (for live price + technical signals). Multiple folios of the same
      fund are merged into one holding per proxy ticker.
    - Debt MF holdings are skipped (not relevant for equity portfolio analysis).

    Returns an ImportReport with details of what was imported.
    """
    report = ImportReport(
        stocks=[],
        mfs=[],
        synced_tickers=[],
        skipped_isins=[],
        mf_proxies_added=[],
        mf_debt_skipped=[],
        errors=[],
    )

    folder_path = Path(folder)
    if not folder_path.exists():
        report.errors.append(f"Folder '{folder}' not found.")
        return report

    xlsx_files = sorted(folder_path.glob("*.xlsx"))
    if not xlsx_files:
        report.errors.append(f"No .xlsx files found in '{folder}'.")
        return report

    today = __import__("datetime").date.today().strftime("%Y-%m-%d")

    for fp in xlsx_files:
        ftype = detect_file_type(fp)

        if ftype == "stock":
            try:
                stocks = parse_stock_holdings(fp)
                report.stocks.extend(stocks)
                for s in stocks:
                    if s.ticker:
                        memory.add_holding(
                            ticker=s.ticker,
                            entry_price=round(s.avg_buy_price, 2),
                            entry_date=today,
                            current_price=round(s.current_price, 2),
                            unrealized_pnl_pct=round(s.unrealized_pnl_pct, 2),
                            thesis=(
                                f"{s.name} ({s.isin}) | Qty: {s.quantity:.0f} units | "
                                f"Imported from broker statement"
                            ),
                            exit_condition=(
                                "Review when macro regime changes or technicals break support."
                            ),
                            quantity=round(s.quantity, 4),
                        )
                        report.synced_tickers.append(s.ticker)
                    else:
                        report.skipped_isins.append(f"{s.isin} ({s.name})")
            except Exception as exc:
                report.errors.append(f"{fp.name} [stock parse error]: {exc}")

        elif ftype == "mf":
            try:
                mfs = parse_mf_holdings(fp)
                report.mfs.extend(mfs)

                # Group equity MFs by proxy ticker (multiple folios -> one holding)
                proxy_groups: Dict[str, Dict] = {}
                for m in mfs:
                    if not m.is_equity:
                        report.mf_debt_skipped.append(m.scheme_name)
                        continue
                    if not m.proxy_ticker:
                        continue  # no proxy available; reported via mfs list
                    if m.proxy_ticker not in proxy_groups:
                        proxy_groups[m.proxy_ticker] = {
                            "schemes": [],
                            "total_invested": 0.0,
                            "total_current": 0.0,
                            "total_units": 0.0,
                            "proxy_note": m.proxy_note,
                            "category": m.category,
                        }
                    g = proxy_groups[m.proxy_ticker]
                    g["schemes"].append(f"{m.scheme_name} ({m.amc})")
                    g["total_invested"] += m.invested_value
                    g["total_current"] += m.current_value
                    g["total_units"] += m.units

                for proxy_ticker, g in proxy_groups.items():
                    invested = g["total_invested"]
                    current = g["total_current"]
                    units = g["total_units"]
                    pnl_pct = ((current - invested) / invested * 100) if invested else 0.0
                    schemes_str = " + ".join(g["schemes"])

                    # Use per-unit NAV as the "price" so the system can reason about cost
                    # basis. The live ETF price will be fetched by yfinance for signals.
                    avg_nav = round(invested / units, 2) if units else 0.0
                    cur_nav = round(current / units, 2) if units else 0.0

                    memory.add_holding(
                        ticker=proxy_ticker,
                        entry_price=avg_nav,
                        entry_date=today,
                        current_price=cur_nav,
                        unrealized_pnl_pct=round(pnl_pct, 2),
                        thesis=(
                            f"MF holding (proxy ETF for signals) | {g['category']} | "
                            f"Invested Rs.{invested:,.0f} -> Current Rs.{current:,.0f} "
                            f"({pnl_pct:+.1f}%) | Schemes: {schemes_str}"
                        ),
                        exit_condition=(
                            f"Exit if {g['category']} macro headwind persists OR "
                            f"index breaks key support. Review XIRR vs FD rate annually."
                        ),
                        quantity=round(units, 4),
                    )
                    report.synced_tickers.append(proxy_ticker)
                    report.mf_proxies_added.append(
                        f"{proxy_ticker} -> {schemes_str}"
                    )

            except Exception as exc:
                report.errors.append(f"{fp.name} [MF parse error]: {exc}")

        else:
            report.errors.append(
                f"{fp.name}: Could not detect file type (expected stock broker "
                f"statement or MF consolidated statement)."
            )

    return report


# ---------------------------------------------------------------------------
# CLI-style pretty print
# ---------------------------------------------------------------------------

def print_import_report(report: ImportReport, W: int = 72):
    print(f"\n{'=' * W}")
    print("  HOLDINGS IMPORT REPORT")
    print(f"{'=' * W}")

    # --- Stock holdings ---
    if report.stocks:
        print(f"\n  Stock ETF Holdings ({len(report.stocks)} found):")
        for s in report.stocks:
            status = f"-> {s.ticker}" if s.ticker else "  [ISIN not mapped — skipped]"
            pnl_str = f"{s.unrealized_pnl_pct:+.1f}%"
            print(f"    {s.name[:40]:<40} {status}")
            print(f"    {'':42} Buy Rs.{s.avg_buy_price:,.2f}  Now Rs.{s.current_price:,.2f}  P&L {pnl_str}")
    else:
        print("\n  No stock ETF holdings found.")

    # --- MF holdings ---
    if report.mfs:
        print(f"\n  Mutual Fund Holdings ({len(report.mfs)} found):")
        for m in report.mfs:
            if not m.is_equity:
                status = "[DEBT — skipped]"
            elif m.proxy_ticker:
                status = f"-> {m.proxy_ticker} (proxy)"
            else:
                status = "[equity, no proxy mapped]"
            pnl_pct = (
                ((m.current_value - m.invested_value) / m.invested_value * 100)
                if m.invested_value else 0.0
            )
            print(f"    {m.scheme_name[:45]:<45} {status}")
            print(f"    {'':47} Inv Rs.{m.invested_value:,.0f}  Now Rs.{m.current_value:,.0f}  "
                  f"P&L {pnl_pct:+.1f}%  XIRR {m.xirr:.1f}%")
    else:
        print("\n  No mutual fund holdings found.")

    # --- Summary ---
    print(f"\n  {'-' * 50}")
    if report.synced_tickers:
        print(f"  [OK] Added to portfolio:  {', '.join(report.synced_tickers)}")
    if report.mf_proxies_added:
        print(f"  [OK] MF proxy holdings:   {len(report.mf_proxies_added)} added to portfolio")
        for item in report.mf_proxies_added:
            print(f"       {item}")
    if report.skipped_isins:
        print(f"  [!]  ISIN not in map ({len(report.skipped_isins)}) — add to ISIN_TO_TICKER:")
        for s in report.skipped_isins:
            print(f"       {s}")
    if report.mf_debt_skipped:
        print(f"  [i]  Debt MFs skipped:    {', '.join(report.mf_debt_skipped)}")
    if report.errors:
        print(f"\n  Errors:")
        for e in report.errors:
            print(f"    [ERR] {e}")

    print(f"\n  Next step:  python run.py show          (verify portfolio)")
    print(f"  Next step:  python run.py portfolio     (run full portfolio analysis)")
    print(f"{'=' * W}\n")
